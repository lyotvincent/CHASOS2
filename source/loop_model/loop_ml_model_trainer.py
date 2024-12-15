'''
train program in py3 for Lollipop, come from Lollipop's train_models.py
'''

import sys, os, json, math
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, HistGradientBoostingClassifier
from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import accuracy_score, roc_auc_score, average_precision_score
# import xgboost as xgb
# from xgboost import XGBClassifier
import joblib
from openpyxl import Workbook, load_workbook
import warnings
from io import BytesIO
from source.parameters import ROOT_DIR, CELL_LINE_LIST, LOOP_MODEL_PATH


def train_model(max_depth=None, max_features=None, n_estimators=None):
    '''
    features info:
    0   chrom
    1   start1
    2   start2
    3   respons
    4   length                  1
    5   motif_pattern           2
    6   avg_motif_strength      3
    7   std_motif_strength      4
    8   avg_conservation        5
    9   std_conservation        6
    10  avg_HS                  7
    11  std_HS                  8
    12  HS_in-between           9
    13  HS_left                 10
    14  HS_right                11
    15  avg_anchor_score        12
    16  std_anchor_score        13
    17  anchor_score1           14
    18  anchor_score2           15
    19  avg_ocr_score           16
    20  std_ocr_score           17
    21  ocr_score1              18
    22  ocr_score2              19
    '''

    # print('Reading in features...')
    # "AG04450", "BJ", "Caco-2", "GM12878", "HCT116", "IMR-90", "K562", "MCF-7"
    data_AG04450 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/AG04450_8cellline_bychr.loop", sep='\t')
    data_BJ = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/BJ_8cellline_bychr.loop", sep='\t')
    data_GM12878 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/GM12878_8cellline_bychr.loop", sep='\t')
    data_IMR90 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/IMR-90_8cellline_bychr.loop", sep='\t')
    # cancer: ["Caco-2", "HCT116", "K562", "MCF-7"]
    data_Caco2 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/Caco-2_8cellline_bychr.loop", sep='\t')
    data_HCT116 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/HCT116_8cellline_bychr.loop", sep='\t')
    data_K562 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/K562_8cellline_bychr.loop", sep='\t')
    data_MCF7 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/MCF-7_8cellline_bychr.loop", sep='\t')
    # data = pd.concat([data_GM12878, data_HCT116, data_IMR90, data_K562], axis=0, ignore_index=True)
    # data = pd.concat([data_AG04450, data_BJ, data_Caco2, data_GM12878, data_HCT116, data_IMR90, data_K562, data_MCF7], axis=0, ignore_index=True)

    # features = data.columns.values[4:]
    # print(features.shape)

    # Evaluating whether the data are balanced..
    # num_pos = data[data['response'] == 1].shape[0]
    # num_neg = data[data['response'] == 0].shape[0]
    # n2p = float(num_neg)/float(num_pos)
    # print('Negative samples were '+str(n2p)+' times more than the positive loops...')

    # val_chr_list = ["chr22", "chrX"]
    # train_data = data[data['chrom'].isin(val_chr_list) == False]
    # val_data = data[data['chrom'].isin(val_chr_list) == True]
    # print(f"train_data.shape: {train_data.shape}")
    # print(f"val_data.shape: {val_data.shape}")

    # divide data into train and val by cell line
    train_data = pd.concat([data_AG04450, data_BJ, data_GM12878, data_IMR90], axis=0, ignore_index=True)
    val_data = pd.concat([data_Caco2, data_HCT116, data_K562, data_MCF7], axis=0, ignore_index=True)

    # num_pos = train_data[train_data['response'] == 1].shape[0]
    # num_neg = train_data[train_data['response'] == 0].shape[0]
    # n2p = float(num_neg)/float(num_pos)

    ## * all data
    # train_X = train_data.iloc[:,4:]
    # val_X = val_data.iloc[:,4:]
    ## * no functional genome data (base)
    # train_X = train_data.iloc[:,4:10]
    # val_X = val_data.iloc[:,4:10]
    ## * add functional genome data (base+functional)
    train_X = train_data.iloc[:,4:15]
    val_X = val_data.iloc[:,4:15]
    ## * add pretrained data (base+myDL)
    # train_X = train_data.iloc[:,[4,5,6,7,8,9,15,16,17,18,19,20,21,22]]
    # val_X = val_data.iloc[:,[4,5,6,7,8,9,15,16,17,18,19,20,21,22]]
    ## * base + myDLOCR
    # train_X = train_data.iloc[:,[4,5,6,7,8,9,19,20,21,22]]
    # val_X = val_data.iloc[:,[4,5,6,7,8,9,19,20,21,22]]
    ## * base + myDLAnchor
    # train_X = train_data.iloc[:,[4,5,6,7,8,9,15,16,17,18]]
    # val_X = val_data.iloc[:,[4,5,6,7,8,9,15,16,17,18]]
    ## * base + funtional + myDLOCR
    # train_X = train_data.iloc[:,[4,5,6,7,8,9,10,11,12,13,14,19,20,21,22]]
    # val_X = val_data.iloc[:,[4,5,6,7,8,9,10,11,12,13,14,19,20,21,22]]
    ## * base + funtional + myDLAnchor
    # train_X = train_data.iloc[:,[4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]]
    # val_X = val_data.iloc[:,[4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]]

    # labels
    train_Y = np.array(train_data['response'])
    val_Y = np.array(val_data['response'])

    print("train_X.shape: ", train_X.shape)
    print("train_Y.shape: ", train_Y.shape)
    print("val_X.shape: ", val_X.shape)
    print("val_Y.shape: ", val_Y.shape)

    # * Lollipop: Use Random Forest classifier to predict interactions
    # run this model with their data: 0.9094202898550725
    # model = RandomForestClassifier(max_depth = 36,max_features=18, n_estimators=100, random_state = np.random.RandomState(0), n_jobs=-1, class_weight={0:1,1:n2p}) 

    # * CharID-Loop: Use Gradient Boosting Classifier to predict interactions
    # run this model with their data: 0.9202898550724637
    # model = GradientBoostingClassifier(n_estimators=235, random_state=10)

    # * my model
    if max_depth == None or max_features == None or n_estimators == None:
        # * base+functional
        # model = GradientBoostingClassifier(max_depth=3, max_features=5, n_estimators=96, random_state=np.random.RandomState(9523))
        # * base+myDL
        # model = GradientBoostingClassifier(max_depth=9, max_features=11, n_estimators=96, random_state=np.random.RandomState(9523))
        model = GradientBoostingClassifier(max_depth=3, max_features=5, n_estimators=256, random_state=np.random.RandomState(9523))
        # model = HistGradientBoostingClassifier(max_depth=3, random_state=np.random.RandomState(9523))
        # model = RandomForestClassifier(max_depth=7, max_features=7, n_estimators=32, n_jobs=-1, random_state=np.random.RandomState(9523))
        # model = XGBClassifier(max_depth=3, n_estimators=128, tree_method='gpu_hist', gpu_id=0, seed=9523)
    else:
        model = GradientBoostingClassifier(max_depth=max_depth, max_features=max_features, n_estimators=n_estimators, random_state=np.random.RandomState(9523))
        # model = RandomForestClassifier(max_depth=max_depth, max_features=max_features, n_estimators=n_estimators, n_jobs=-1, random_state=np.random.RandomState(9523))
        # model = XGBClassifier(max_depth=max_depth, n_estimators=n_estimators, tree_method='gpu_hist', gpu_id=0, seed=9523)

    # print('Training model...')
    EPOCH = 10
    best_epoch = 0
    best_score = 0
    for i in range(EPOCH):
        model.fit(train_X, train_Y)
        r_score = model.score(val_X, val_Y)
        if r_score > best_score:
            best_epoch = i
            best_score = r_score
            #lr是一个LogisticRegression模型
            joblib.dump(model, LOOP_MODEL_PATH+'/model/RandomForestClassifier_mine_basefunctional_8cellline_byhealth.joblib')
        print(i, r_score)
    print('best score: ', best_epoch, best_score)
    best_model = joblib.load(LOOP_MODEL_PATH+'/model/RandomForestClassifier_mine_basefunctional_8cellline_byhealth.joblib')
    r_score = best_model.score(val_X, val_Y)
    print('best acc2: ', r_score)
    print(model.feature_importances_)
    # val_X = val_X.to_numpy() # DataFrame -> ndarray, test np data is useful?
    # t = best_model.predict(val_X)
    # pred_Y = list(map(int, t >= 0.5))
    # acc = accuracy_score(val_Y, pred_Y)
    # print('best score3: ', acc)

    # auc
    predictions = best_model.predict_proba(val_X)
    predictions = predictions[:, 1]
    roc_auc = roc_auc_score(val_Y, predictions)
    prc_auc = average_precision_score(val_Y, predictions)
    print('roc_auc: ', roc_auc)
    print('prc_auc: ', prc_auc)

    # trees = best_model.estimators_
    # booster = best_model.get_booster()
    # print("trees num: ", booster.num_boosted_rounds())
    # warnings.filterwarnings("ignore")
    # for i in range(32):
    #     t = trees[i][0].predict(val_X)
    #     print(t.shape)
    #     pred_Y = list(map(int, t >= 0.5))
    #     acc = accuracy_score(val_Y, pred_Y)
    #     print(acc)
    # for tree_dump in booster:
    #     # booster = xgb.Booster(model_file=None)
    #     # booster.load_model(BytesIO(tree_dump.encode()))
    #     dtrain = xgb.DMatrix(val_X) # , feature_names=val_X.columns
    #     # 单独对输入数据进行预测
    #     pred_Y = tree_dump.predict(dtrain)
    #     # print(type(pred_Y), pred_Y.shape)
    #     pred_Y = list(map(int, pred_Y >= 0.5))
    #     acc = accuracy_score(val_Y, pred_Y)
    #     print(acc)
    return best_score

def draw_gridsearch():
    from matplotlib import pyplot as plt
    wb = load_workbook(ROOT_DIR+'/temp/loop_model_record.xlsx', data_only=True)
    sheet1 = wb['base+functional']
    rows = sheet1.rows
    rows = list(rows)[1:]
    x = list() # max_depth
    y = list() # max_features
    z = list() # n_estimators
    acc = list()
    for row in rows:
        x.append(row[0].value)
        y.append(row[1].value)
        z.append(row[2].value)
        acc.append(row[3].value)
    min_acc, max_acc = min(acc), max(acc)
    min_max_normalization = [(i-min_acc)/(max_acc-min_acc+0.000001) for i in acc]
    colorr = ['#277DA1','#577590','#4D908E','#43AA8B','#90BE6D','#F9c74F','#F9844A','#F8961E','#F3722c','#F94144']
    # acc_c = [(i, 0, 0) for i in min_max_normalization]
    # acc_c = list()
    # for i in min_max_normalization:
    #     print(i, math.floor(i*10))
    #     acc_c.append(colorr[math.floor(i*10)])
    acc_c = [colorr[math.floor(i*10)] for i in min_max_normalization]
    acc_s = [i**3*100 for i in min_max_normalization]

    #定义坐标轴
    fig = plt.figure()
    ax = fig.add_subplot(111, projection="3d") # type: ignore
    # x = [3, 5, 8, 12, 16, 8, 8, 8, 8, 16]
    # y = [19, 19, 19, 19, 19, 19, 19, 19, 19, 19]
    # z = [128, 128, 128, 128, 128, 256, 512, 64, 32, 32]
    # acc = [0.9130434782608695, 0.9166666666666666, 0.927536231884058, 0.9202898550724637, 0.8840579710144928, 0.9239130434782609, 0.92391304347826, 0.927536231884058, 0.92391304347826, 0.880434782608695]
    # acc_c = [((i-0.85)*10, 0, 0) for i in acc]
    # acc_s = [(i-0.85)*10*200 for i in acc]
    for i in range(len(x)):
        ax.scatter(x[i], y[i], zs=z[i], zdir="z", color=acc_c[i], s=acc_s[i])
        # ax.scatter(x[i], y[i], zs=z[i], zdir="z", color=acc_c[i])
    plt.xticks([3, 5, 7, 9, 11, 13])
    plt.yticks([3, 5, 7, 9, 11])
    plt.tight_layout()
    plt.show()

def gridsearch():
    '''搜索整个空间，探寻使用我的深度学习得到的特征和原生测序特征的最佳参数'''
    # 创建一个Workbook对象
    wb = Workbook()
    # 如果不指定sheet索引和表名，默认在第二张表位置新建表名sheet1
    # wb.create_sheet(index=1, title="sheet2") 

    # 获取当前活跃的sheet，默认为第一张sheet
    ws = wb.active
    ws['A1'] = 'max_depth' # type: ignore
    ws['B1'] = 'max_features' # type: ignore
    ws['C1'] = 'n_estimators' # type: ignore
    ws['D1'] = 'acc' # type: ignore
    # base
    # max_depth_list = [3, 5, 7, 9, 11, 13]
    # max_features_list = [3, 5, 7, 9]
    # n_estimators_list = [32*i for i in range(1, 17)]
    # base+functional
    # max_depth_list = [3, 5, 7, 9, 11, 13]
    # max_features_list = [3, 5, 7, 9, 11]
    # n_estimators_list = [32*i for i in range(1, 17)]
    # # base+myDL
    max_depth_list = [3, 5, 7, 9, 11, 13]
    max_features_list = [3, 5, 7, 9, 11, 13, 14]
    max_features_list = [0] # for xgb
    n_estimators_list = [32*i for i in range(1, 17)]
    max_acc = 0
    for a in max_depth_list:
        for b in max_features_list:
            for c in n_estimators_list:
                score = train_model(a, b, c)
                row = [a, b, c, score]
                if score > max_acc:
                    max_acc = score
                print(row, max_acc)
                ws.append(row) # type: ignore

    wb.save('base+myDL.xlsx')

def excel2json():
    wb = load_workbook(ROOT_DIR+'/temp/loop_model_record.xlsx', data_only=True)
    # sheets = wb.worksheets   # 获取当前所有的sheet
    # print(sheets)

    # 获取第一张sheet
    # sheet1 = sheets[0]
    # sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
    sheet1 = wb['base+functional']
    # print(sheet1)
    rows = sheet1.rows
    # 通过Cell对象读取
    # cell_11 = sheet1.cell(1,1).value
    # print(cell_11)

    json_list = list()
    for row in rows:
        # print(row)
        # row_val = [col.value for col in row]
        max_depth = row[0].value
        max_features = row[1].value
        n_estimators = row[2].value
        acc = row[3].value
        min_max_normalization = row[4].value
        symbolsize = 0.1
        # print(row_val)
        json_list.append([max_depth, max_features, n_estimators, acc, min_max_normalization, symbolsize])
    with open(ROOT_DIR+'/figure/loop_model_record.json', 'w') as f:
        json.dump(json_list, f)

def evaluate_model():
    data_AG04450 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/AG04450.loop", sep='\t')
    data_BJ = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/BJ.loop", sep='\t')
    data_GM12878 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/GM12878.loop", sep='\t')
    data_IMR90 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/IMR-90.loop", sep='\t')
    # cancer: ["Caco-2", "HCT116", "K562", "MCF-7"]
    data_Caco2 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/Caco-2.loop", sep='\t')
    data_HCT116 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/HCT116.loop", sep='\t')
    data_K562 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/K562.loop", sep='\t')
    data_MCF7 = pd.read_table(LOOP_MODEL_PATH+"/ml_training_data/MCF-7.loop", sep='\t')
    train_data = pd.concat([data_AG04450, data_BJ, data_GM12878, data_IMR90], axis=0, ignore_index=True)
    val_data = pd.concat([data_Caco2, data_HCT116, data_K562, data_MCF7], axis=0, ignore_index=True)

    ## * all data (all)
    # train_X = train_data.iloc[:,4:]
    # val_X = val_data.iloc[:,4:]
    ## * no functional genome data (base)
    # train_X = train_data.iloc[:,4:10]
    # val_X = val_data.iloc[:,4:10]
    ## * add functional genome data (base+functional)
    # train_X = train_data.iloc[:,4:15]
    # val_X = val_data.iloc[:,4:15]
    ## * add pretrained data (base+myDL)
    train_X = train_data.iloc[:,[4,5,6,7,8,9,15,16,17,18,19,20,21,22]]
    val_X = val_data.iloc[:,[4,5,6,7,8,9,15,16,17,18,19,20,21,22]]

    # labels
    train_Y = np.array(train_data['response'])
    val_Y = np.array(val_data['response'])

    best_model = joblib.load(LOOP_MODEL_PATH+'/model/GradientBoostingClassifier_8cellline_healthcancer.joblib')
    r_score = best_model.score(val_X, val_Y)
    print('best score2: ', r_score)

    # auc
    predictions = best_model.predict_proba(val_X)
    predictions = predictions[:, 1]
    roc_auc = roc_auc_score(val_Y, predictions)
    prc_auc = average_precision_score(val_Y, predictions)
    print('roc_auc: ', roc_auc)
    print('prc_auc: ', prc_auc)


if __name__ == "__main__":
    train_model()
    # draw_gridsearch()
    # gridsearch()
    # excel2json()
    # evaluate_model()

